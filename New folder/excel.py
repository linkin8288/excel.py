from openpyxl import Workbook
wb = Workbook() 

# 第三方套件解說:
# type: Workbook 這是Workbook type，套件裡面自己發明的type
# x = 5  data type: int，python裡面每個東西都是物件
# 要如何發明自己的型別type呢? 要寫class
# 先建立物件 wb = Workbook()，再使用那個物件的功能，再儲存下來


# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42
ws['B1'] = 'Allen'

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")